import argparse
import json
import unicodedata
from pathlib import Path
from urllib import error, request

from openpyxl import load_workbook


def to_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value):
    text = unicodedata.normalize("NFKD", to_text(value))
    text = text.encode("ascii", "ignore").decode("ascii")
    return text.strip().lower()


def to_number(value, default=0):
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace("Q", "").replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return default


def to_int(value, default=0):
    return int(round(to_number(value, default)))


def normalize_row(row):
    sku = to_text(row.get("sku"))
    category = to_text(row.get("categoria"))
    brand = to_text(row.get("marca"))
    model = to_text(row.get("modelo"))
    series = to_text(row.get("serie/version"))
    quality = to_text(row.get("calidad"))
    color = to_text(row.get("color"))
    observations = to_text(row.get("observaciones"))
    warranty = to_text(row.get("aplica garantia"))

    if not sku or not category or not brand:
        return None

    purchase_price = to_number(row.get("costo compra unidad"))
    if purchase_price <= 0:
        purchase_price = to_number(row.get("costo compra mayorista"))

    public_price = to_number(row.get("precio cliente final"))
    tech_price = to_number(row.get("precio tecnico/codigo"), None)

    compatible = " ".join(part for part in [brand, model, series] if part)
    warranty_policy = ""
    if warranty:
        warranty_policy = "Aplica garantia: {}".format(warranty.upper())
    if observations:
        warranty_policy = "{}. {}".format(warranty_policy, observations).strip(". ").strip()

    return {
        "internalCode": sku,
        "name": " ".join(part for part in [brand, model] if part),
        "category": category,
        "purchasePrice": purchase_price,
        "publicSalePrice": public_price,
        "currentStock": to_int(row.get("stock actual")),
        "minimumStock": max(0, to_int(row.get("stock minimo"), 2)),
        "description": observations or None,
        "compatibleWith": compatible or None,
        "brand": brand or None,
        "model": model or None,
        "series": series or None,
        "quality": quality or None,
        "color": color or None,
        "technicianSalePrice": tech_price if tech_price and tech_price > 0 else None,
        "location": None,
        "supplier": None,
        "warrantyPolicy": warranty_policy or None,
    }


def chunks(items, size):
    for index in range(0, len(items), size):
        yield items[index : index + size]


def api_call(url, method="GET", payload=None, token=None):
    headers = {}
    data = None
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    if token:
        headers["Authorization"] = "Bearer {}".format(token)
    req = request.Request(url, data=data, headers=headers, method=method)
    with request.urlopen(req) as response:
        return json.loads(response.read().decode("utf-8"))


def main():
    parser = argparse.ArgumentParser(description="Importa inventario profesional al backend del taller.")
    parser.add_argument("--file", default=str(Path.home() / "Downloads" / "inventario_repuestos_profesional.xlsx"))
    parser.add_argument("--api", default="http://127.0.0.1:3000/api")
    parser.add_argument("--user", default="admin")
    parser.add_argument("--password", default="Admin123*")
    parser.add_argument("--sheet", default="01_Inventario")
    parser.add_argument("--batch-size", type=int, default=40)
    args = parser.parse_args()

    workbook = load_workbook(args.file, data_only=True)
    worksheet = workbook[args.sheet]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [normalize_key(value) for value in rows[0]]

    items = []
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        normalized = normalize_row(row)
        if normalized and normalized["publicSalePrice"] > 0:
            items.append(normalized)

    if not items:
        print("No se encontraron repuestos validos para importar.")
        return 1

    login = api_call(
        "{}/auth/login".format(args.api),
        method="POST",
        payload={"username": args.user, "password": args.password},
    )
    token = login["accessToken"]

    created = 0
    updated = 0
    for batch in chunks(items, args.batch_size):
        result = api_call(
            "{}/spare-parts/bulk-import".format(args.api),
            method="POST",
            payload={"items": batch},
            token=token,
        )
        created += int(result.get("created", 0))
        updated += int(result.get("updated", 0))

    print("Importacion completada.")
    print("Archivo:", args.file)
    print("Repuestos procesados:", len(items))
    print("Creados:", created)
    print("Actualizados:", updated)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        print("Error HTTP {}: {}".format(exc.code, body))
        raise SystemExit(1)
    except Exception as exc:
        print("Error:", exc)
        raise SystemExit(1)
